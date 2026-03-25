from fastapi import APIRouter, UploadFile, HTTPException, Depends, File, Query
import pandas as pd
from fastapi.responses import StreamingResponse
from typing import Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, asc, desc, func
from dependencies.permission import PermissionChecker
from models.cost import CostImport, CostHRImport
from database import get_db_cost
import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from io import BytesIO
from helper import getStore, getHRStore
from graphqlschema.department import getDepartments
from graphqlschema.schema import DepartmentSearchParameter
import re

router = APIRouter(prefix="/costs", tags=["Cost"])
BASE_DIR = Path(__file__).parent.parent  # 退一级
UPLOAD_DIR = BASE_DIR / "uploads" / "costs"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

def parse_year_month(year_str: str, month_str: str) -> str:
    """将 year 和 month 字符串合并为 YYYY-MM 格式，支持 '07' 或 'Jul' """
    month_map = {
        'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
        'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
        'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
    }

    if not re.fullmatch(r"\d{4}", year_str):
        raise ValueError(f"年份格式错误: {year_str}")

    month_str = month_str.strip().capitalize()

    if month_str.isdigit():
        if 1 <= int(month_str) <= 12:
            month_num = month_str.zfill(2)
        else:
            raise ValueError(f"无效数字月份: {month_str}")
    elif month_str in month_map:
        month_num = month_map[month_str]
    else:
        raise ValueError(f"无效英文月份缩写: {month_str}")

    return f"{year_str}-{month_num}"

@router.get("/template", summary="下载成本录入模板（含限制）")
async def download_cost_template(
    user=Depends(PermissionChecker(required_roles=["cost:download"])),
    db: AsyncSession = Depends(get_db_cost),
):
    valid_stores = getStore()
    valid_departments = [item.name["en_us"] for item in getDepartments(None).departments]

    # ✅ 创建工作簿
    wb = Workbook()

    # ✅ 先创建 Data sheet
    ws_data = wb.active
    ws_data.title = "Data"
    ws_data.append(["store", "department", "year", "month", "cost"])

    # ✅ 创建 Reference sheet
    ws_ref = wb.create_sheet(title="Reference")

    # 写入门店
    ws_ref.append(["Store"])
    for store in valid_stores:
        ws_ref.append([store])
    ws_ref.append([])  # 空行
    ws_ref.append(["Department"])
    for dept in valid_departments:
        ws_ref.append([dept])
    ws_ref.column_dimensions['A'].hidden = True

    # ✅ 定义数据有效性范围
    # store 开始范围：Reference!A2:A(n+1)
    store_start_row = 2
    store_end_row = store_start_row + len(valid_stores) - 1
    #store_range = f"Reference!A{store_start_row}:A{store_end_row}"
    # department 开始范围：Reference!A(len(valid_stores)+4) 到末尾
    dept_start_row = len(valid_stores) + 4
    dept_end_row = dept_start_row + len(valid_departments) - 1
    #dept_range = f"Reference!A{dept_start_row}:A{dept_end_row}"
    # ✅ 定义数据有效性范围 (绝对引用)
    store_range = f"'Reference'!$A$2:$A${store_end_row}"
    dept_range = f"'Reference'!$A${dept_start_row}:$A${dept_end_row}"


    # ✅ 创建数据验证对象
    store_validation = DataValidation(
        type="list",
        formula1=store_range,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle='无效门店',
        error='请从下拉列表中选择有效门店',
    )
    dept_validation = DataValidation(
        type="list",
        formula1=dept_range,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle='无效部门',
        error='请从下拉列表中选择有效部门',
    )
    ws_data.add_data_validation(store_validation)
    ws_data.add_data_validation(dept_validation)

    # ✅ 一次性添加整列范围
    store_validation.add(f"A2:A500")
    dept_validation.add(f"B2:B500")
    # ✅ 把验证应用到每一行
    # for row in range(2, 500):  # 500 行范围
    #     store_validation.add(ws_data[f"A{row}"])
    #     dept_validation.add(ws_data[f"B{row}"])

    # ✅ 输出成字节流
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cost_template.xlsx"},
    )

@router.get("/template/hr", summary="下载HR成本录入模板（含限制）")
async def download_hr_cost_template(
    user=Depends(PermissionChecker(required_roles=["cost:download"])),
    db: AsyncSession = Depends(get_db_cost),
):
    valid_stores = getHRStore()
    
    # 1. 预取所有门店对应的 HR 部门
    store_dept_map = {}
    for store in valid_stores:
        param = DepartmentSearchParameter(HR=True, Store=store)
        try:
            depts_data = getDepartments(param)
            # name 是 JSON 类型 (dict)，取 "en_us"
            dept_names = [item.name["en_us"] for item in depts_data.departments if item.name]
            store_dept_map[store] = dept_names
        except Exception:
            store_dept_map[store] = []

    wb = Workbook()

    # 2. Data Sheet
    ws_data = wb.active
    ws_data.title = "Data"
    ws_data.append(["store", "department", "year", "month", "labor_cost", "other_cost"])

    # 3. Reference Sheet (用于存放下拉选项源数据)
    ws_ref = wb.create_sheet(title="Reference")

    # 3.1 写入门店列表到 A 列
    ws_ref.append(["Store"])
    for store in valid_stores:
        ws_ref.append([store])
    # 创建 StoreList 命名范围 (使用 DefinedName)
    dn = DefinedName("StoreList", attr_text=f"'Reference'!$A$2:$A${len(valid_stores) + 1}")
    wb.defined_names.add(dn)

    # 3.2 写入各门店部门到后续列 (C列开始)
    col_idx = 3
    for store in valid_stores:
        depts = store_dept_map.get(store, [])
        if not depts:
            continue
        
        # 将该店的部门写入一列
        ws_ref.cell(row=1, column=col_idx, value=store)
        for i, dept in enumerate(depts):
            ws_ref.cell(row=i+2, column=col_idx, value=dept)
        
        # 创建命名范围：HR_STORE_SanitizedName
        # 必须加上前缀避免与单元格地址(如B1)冲突，并处理特殊字符
        sanitized_store = re.sub(r'[^a-zA-Z0-9]', '_', store)
        range_name = f"HR_STORE_{sanitized_store}"
        
        col_letter = get_column_letter(col_idx)
        dn = DefinedName(range_name, attr_text=f"'Reference'!${col_letter}$2:${col_letter}${len(depts)+1}")
        wb.defined_names.add(dn)
        col_idx += 1

    ws_ref.sheet_state = "hidden"

    # 4. 数据验证
    # Store 列使用 StoreList
    store_validation = DataValidation(type="list", formula1="=StoreList", allow_blank=True, showErrorMessage=True, errorTitle='无效门店', error='请选择有效门店')
    ws_data.add_data_validation(store_validation)
    store_validation.add("A2:A500")

    # Department 列使用 INDIRECT 引用对应门店的 Named Range
    dept_validation = DataValidation(type="list", formula1='=INDIRECT("HR_STORE_"&SUBSTITUTE(A2," ","_"))', allow_blank=True, showErrorMessage=True, errorTitle='无效部门', error='请选择该门店下的有效部门')
    ws_data.add_data_validation(dept_validation)
    dept_validation.add("B2:B500")

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=hr_cost_template.xlsx"},
    )

@router.post("/upload/hr", summary="上传hr成本 Excel 文件")
async def upload_hr_cost_xlsx(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db_cost),
    user=Depends(PermissionChecker(required_roles=["cost:insert"])),
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 xlsx 文件")

    contents = await file.read()

    try:
        wb = load_workbook(BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取文件: {e}")

    if "Data" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="缺少 Data sheet")
    ws = wb["Data"]

    header = [str(cell.value).strip().lower() for cell in next(ws.iter_rows(max_row=1))]
    # other_cost 是可选的，但在新表中是必须字段(允许为空)，这里检查excel是否有这一列
    required_cols = {"store", "department", "year", "month", "labor_cost"}
    if not required_cols.issubset(header):
        raise HTTPException(status_code=400, detail=f"缺少必须列: {required_cols}")

    col_idx = {name: header.index(name) for name in header}
    valid_stores = getHRStore()

    # 缓存: store -> { full_name -> {id, short_name} }
    store_dept_map_cache = {}

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        store = str(row[col_idx["store"]]).strip() if row[col_idx["store"]] else None
        dept_full = str(row[col_idx["department"]]).strip() if row[col_idx["department"]] else None
        year_cell = row[col_idx["year"]]
        month_cell = row[col_idx["month"]]
        cost_val = row[col_idx["labor_cost"]]
        other_cost_val = row[col_idx["other_cost"]] if "other_cost" in col_idx and row[col_idx["other_cost"]] is not None else 0

        if not store and not dept_full and year_cell is None:
            raise HTTPException(status_code=400, detail=f"第{i}行: 请填写 store, department, year")

        if not store or not dept_full or year_cell is None or month_cell is None or cost_val is None:
            raise HTTPException(status_code=400, detail=f"第{i}行: 请填写完整信息")

        year = str(year_cell).strip()
        month = str(month_cell).strip()

        try:
            month_str = parse_year_month(year, month)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=f"第{i}行: {e}")

        if store not in valid_stores:
            raise HTTPException(status_code=400, detail=f"第{i}行: {store} 不是有效HR门店")

        # 加载并缓存该门店的 HR 部门信息
        if store not in store_dept_map_cache:
            param = DepartmentSearchParameter(HR=True, Store=store)
            try:
                depts_data = getDepartments(param)
                mapping = {}
                for d in depts_data.departments:
                    fname = d.name.get("en_us", "")
                    if fname:
                        sname = fname.split("/")[-1] # 取最后一段作为简单名称
                        mapping[fname] = {"id": d.id, "name": sname}
                store_dept_map_cache[store] = mapping
            except Exception:
                store_dept_map_cache[store] = {}

        if dept_full not in store_dept_map_cache[store]:
             raise HTTPException(status_code=400, detail=f"第{i}行: {dept_full} 不是门店 {store} 的有效部门")

        dept_info = store_dept_map_cache[store][dept_full]
        dept_id = dept_info["id"]
        dept_simple_name = dept_info["name"]

        try:
            cost = float(cost_val)
            other_cost = float(other_cost_val)
        except Exception:
            raise HTTPException(status_code=400, detail=f"第{i}行: cost/other_cost 数值无效")

        total_cost = cost + other_cost

        # 检查是否存在记录 (按 store, department_id, month)
        stmt = select(CostHRImport).where(
            CostHRImport.store == store,
            CostHRImport.department_id == str(dept_id),
            CostHRImport.month == month_str,
        )
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()

        if existing:
            existing.cost = cost
            existing.other_cost = other_cost
            existing.total_cost = total_cost
            existing.department = dept_simple_name
            existing.department_full_name = dept_full
            existing.updated_at = datetime.datetime.now()
        else:
            new_record = CostHRImport(
                store=store,
                department=dept_simple_name,
                month=month_str,
                cost=cost,
                created_at=datetime.datetime.now(),
                updated_at=datetime.datetime.now(),
                department_full_name=dept_full,
                department_id=str(dept_id),
                other_cost=other_cost,
                total_cost=total_cost,
                creator_id=str(user.id)
            )
            db.add(new_record)

    await db.commit()

    filename = f"HR_{datetime.datetime.now():%Y%m%d%H%M%S}_{file.filename}"
    file_path = UPLOAD_DIR / filename
    file_path.write_bytes(contents)

    return {"message": "HR成本导入完成并已更新现有记录", "saved_file": str(file_path)}

@router.post("/upload", summary="上传成本 Excel 文件")
async def upload_cost_xlsx(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db_cost),
    user=Depends(PermissionChecker(required_roles=["cost:insert"])),
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 xlsx 文件")

    contents = await file.read()

    try:
        wb = load_workbook(BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取文件: {e}")

    if "Data" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="缺少 Data sheet")
    ws = wb["Data"]

    header = [str(cell.value).strip().lower() for cell in next(ws.iter_rows(max_row=1))]
    required_cols = {"store", "department", "year", "month", "cost"}
    if not required_cols.issubset(header):
        raise HTTPException(status_code=400, detail=f"缺少必须列: {required_cols}")

    col_idx = {name: header.index(name) for name in header}
    valid_stores = getStore()
    valid_departments = [item.name["en_us"] for item in getDepartments(None).departments]

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        store = str(row[col_idx["store"]]).strip() if row[col_idx["store"]] else None
        dept = str(row[col_idx["department"]]).strip() if row[col_idx["department"]] else None
        year_cell = row[col_idx["year"]]
        month_cell = row[col_idx["month"]]
        cost = row[col_idx["cost"]]

        if not store or not dept or year_cell is None or month_cell is None or cost is None:
            continue

        year = str(year_cell).strip()
        month = str(month_cell).strip()

        try:
            month_str = parse_year_month(year, month)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=f"第{i}行: {e}")

        if store not in valid_stores:
            raise HTTPException(status_code=400, detail=f"第{i}行: {store} 不是有效门店")
        if dept not in valid_departments:
            raise HTTPException(status_code=400, detail=f"第{i}行: {dept} 不是有效部门")
        try:
            cost = float(cost)
        except Exception:
            raise HTTPException(status_code=400, detail=f"第{i}行: cost 值无效 '{row[col_idx['cost']]}'")

        stmt = select(CostImport).where(
            CostImport.store == store,
            CostImport.department == dept,
            CostImport.month == month_str,
        )
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()

        if existing:
            existing.cost = cost
            existing.updated_at = datetime.datetime.now()
        else:
            new_record = CostImport(
                store=store,
                department=dept,
                month=month_str,
                cost=cost,
                created_at=datetime.datetime.now(),
                updated_at=datetime.datetime.now(),
            )
            db.add(new_record)

    await db.commit()

    filename = f"{datetime.datetime.now():%Y%m%d%H%M%S}_{file.filename}"
    file_path = UPLOAD_DIR / filename
    file_path.write_bytes(contents)

    return {"message": "导入完成并已更新现有记录", "saved_file": str(file_path)}

@router.get("/list/hr", summary="查询hr成本信息（支持筛选、分页、排序）")
async def list_hr_costs(
    store: Optional[str] = Query(None, description="筛选门店，例如 B1, BVW"),
    department: Optional[str] = Query(None, description="筛选部门，例如 Donation, Grocery"),
    month: Optional[str] = Query(
        None,
        regex=r"^\d{4}-(0[1-9]|1[0-2])$",
        description="筛选年月，格式 YYYY-MM",
    ),
    page: int = Query(1, ge=1, description="页码（从 1 开始）"),
    page_size: int = Query(10, ge=1, le=1000000, description="每页数量（最多100）"),
    sort_by: str = Query(
        "month",
        description="排序字段，可选 id, store, department, month, labor_cost, other_cost, total_cost, created_at, updated_at",
    ),
    sort_order: str = Query(
        "asc",
        regex="^(asc|desc)$",
        description="排序方向 asc 或 desc",
    ),
    db: AsyncSession = Depends(get_db_cost),
    user=Depends(PermissionChecker(required_roles=["cost:search"])),
):
    # 先构造筛选条件，用于数据查询
    data_stmt = select(CostHRImport)
    if store:
        data_stmt = data_stmt.where(CostHRImport.store == store)
    if department:
        data_stmt = data_stmt.where(CostHRImport.department_full_name == department)
    if month:
        data_stmt = data_stmt.where(CostHRImport.month == month)

    # 构造计数查询，只统计符合条件的总行数
    count_stmt = select(func.count()).select_from(CostHRImport)
    if store:
        count_stmt = count_stmt.where(CostHRImport.store == store)
    if department:
        count_stmt = count_stmt.where(CostHRImport.department_full_name == department)
    if month:
        count_stmt = count_stmt.where(CostHRImport.month == month)

    count_result = await db.execute(count_stmt)
    total = count_result.scalar() or 0

    # 排序
    if sort_by == "labor_cost":
        sort_column = CostHRImport.cost
    else:
        sort_column = getattr(CostHRImport, sort_by)
    data_stmt = data_stmt.order_by(
        asc(sort_column) if sort_order == "asc" else desc(sort_column)
    )

    # 分页
    offset = (page - 1) * page_size
    data_stmt = data_stmt.offset(offset).limit(page_size)

    result = await db.execute(data_stmt)
    records = result.scalars().all()

    total_pages = (total + page_size - 1) // page_size

    return {
        "total": total,
        "total_pages": total_pages,
        "current_page": page,
        "page_size": page_size,
        "items": [
            {
                "id": rec.id,
                "store": rec.store,
                "department": rec.department,
                "department_full_name": rec.department_full_name,
                "department_id": rec.department_id,
                "month": rec.month,
                "labor_cost": float(rec.cost or 0),
                "other_cost": float(rec.other_cost or 0),
                "total_cost": float(rec.total_cost or 0),
                "created_at": rec.created_at,
                "updated_at": rec.updated_at,
            }
            for rec in records
        ],
    }


@router.get("/list", summary="查询成本信息（支持筛选、分页、排序）")
async def list_costs(
    store: Optional[str] = Query(None, description="筛选门店，例如 MT, NY"),
    department: Optional[str] = Query(None, description="筛选部门，例如 Donation, Grocery"),
    month: Optional[str] = Query(
        None,
        regex=r"^\d{4}-(0[1-9]|1[0-2])$",
        description="筛选年月，格式 YYYY-MM",
    ),
    page: int = Query(1, ge=1, description="页码（从 1 开始）"),
    page_size: int = Query(10, ge=1, le=1000000, description="每页数量（最多100）"),
    sort_by: str = Query(
        "month",
        description="排序字段，可选 id, store, department, month, cost, created_at, updated_at",
    ),
    sort_order: str = Query(
        "asc",
        regex="^(asc|desc)$",
        description="排序方向 asc 或 desc",
    ),
    db: AsyncSession = Depends(get_db_cost),
    user=Depends(PermissionChecker(required_roles=["cost:search"])),
):
    # 先构造筛选条件，用于数据查询
    data_stmt = select(CostImport)
    if store:
        data_stmt = data_stmt.where(CostImport.store == store)
    if department:
        data_stmt = data_stmt.where(CostImport.department == department)
    if month:
        data_stmt = data_stmt.where(CostImport.month == month)

    # 构造计数查询，只统计符合条件的总行数
    count_stmt = select(func.count()).select_from(CostImport)
    if store:
        count_stmt = count_stmt.where(CostImport.store == store)
    if department:
        count_stmt = count_stmt.where(CostImport.department == department)
    if month:
        count_stmt = count_stmt.where(CostImport.month == month)

    count_result = await db.execute(count_stmt)
    total = count_result.scalar() or 0

    # 排序
    sort_column = getattr(CostImport, sort_by)
    data_stmt = data_stmt.order_by(
        asc(sort_column) if sort_order == "asc" else desc(sort_column)
    )

    # 分页
    offset = (page - 1) * page_size
    data_stmt = data_stmt.offset(offset).limit(page_size)

    result = await db.execute(data_stmt)
    records = result.scalars().all()

    total_pages = (total + page_size - 1) // page_size

    return {
        "total": total,
        "total_pages": total_pages,
        "current_page": page,
        "page_size": page_size,
        "items": [
            {
                "id": rec.id,
                "store": rec.store,
                "department": rec.department,
                "month": rec.month,
                "cost": float(rec.cost),
                "created_at": rec.created_at,
                "updated_at": rec.updated_at,
            }
            for rec in records
        ],
    }